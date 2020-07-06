import os
import sys
from openpyxl import Workbook,load_workbook
import openpyxl

base_path = os.getcwd()
sys.path.append(os.path.dirname(base_path))

def copyexcel(excelpath1=None,excelpath2=None):
    wb2=Workbook()#新建一个excel
    # wb2.save(excelpath2)
    wb1=openpyxl.load_workbook(excelpath1)
    # wb2=openpyxl.load_workbook(excelpath2)
    sheets1=wb1.sheetnames
    sheets2=wb2.sheetnames
    sheet1=wb1[sheets1[0]]
    sheet2=wb2[sheets2[0]]
    max_row=sheet1.max_row
    max_column=sheet1.max_column
    for m in range(1, max_row + 1):
        for n in range(97, 97 + max_column):  # chr(97)='a'
            n = chr(n)  # ASCII字符
            i = '%s%d' % (n, m)  # 单元格编号
            cell1 = sheet1[i].value  # 获取data单元格数据
            sheet2[i].value = cell1  # 赋值到test单元格
    wb2.save(excelpath2)
    wb1.close()
    wb2.close()

class HandExcle:

    def get_excle(self, file_path=None):
        if file_path == None:
            if "/var/lib/jenkins/workspace/" in base_path:
                excel = openpyxl.load_workbook(base_path + '/case/case.xlsx')
            else:
                excel = openpyxl.load_workbook(os.path.dirname(base_path) + '/case/case.xlsx')
        else:
            excel = openpyxl.load_workbook(file_path)

        return excel

    def get_sheet_data(self, index=None):
        '''
        加载所有sheet内容
        '''
        sheet_name = self.get_excle(file_path=os.path.dirname(base_path) + "/report/api_result.xlsx").sheetnames
        if index == None:
            index = 0
        data = self.get_excle(file_path=os.path.dirname(base_path) + "/report/api_result.xlsx")[sheet_name[index]]
        return data

    def get_cell_value(self, row, cols):
        '''
        获取某一个单元格的内容
        '''
        data = self.get_sheet_data().cell(row=row, column=cols).value
        return data

    def get_rows(self):
        '''
        获取行数
        '''
        row = self.get_sheet_data().max_row
        return row

    def get_row_value(self, row):
        '''获取某一行的内容'''
        row_list = []
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list

    def excel_write_data(self, row, cols, value):
        wb = self.get_excle(file_path=os.path.dirname(base_path) + "/report/api_result.xlsx")
        wr = wb.active
        wr.cell(row, cols, value)
        wb.save(os.path.dirname(base_path) + "/report/api_result.xlsx")

    def get_columns_value(self, key=None):
        '''
        获取某一列的数据
        '''
        columns_list = []
        if key == None:
            key = "A"
        columns_list_data = self.get_sheet_data()[key]
        for i in columns_list_data:
            columns_list.append(i.value)
        return columns_list

    def get_rows_number(self, case_id):
        '''
        获取前置条件的用例所在的行号
        '''
        num = 1
        cols_data = self.get_columns_value()
        for col_data in cols_data:
            if case_id == col_data:
                return num
            num = num + 1
        return num

    def \
            get_excel_data(self):
        
        '''
        获取excel里面所有的数据
        '''
        data_list = []
        for i in range(self.get_rows()):
            data_list.append(self.get_row_value(i + 2))
        return data_list

    def get_success_result(self):

        res_list=self.get_columns_value("M")
        if "fail" in res_list:
            print("有失败的")
            return False
        else:
            print("全部成功")
            return True


excel_data = HandExcle()
if __name__ == '__main__':
    # test_excel = os.path.dirname(base_path) + '/case/case.xlsx'
    # report_result = os.path.dirname(base_path) + "/report/api_result.xlsx"
    # hand_excle = HandExcle()
    # print(hand_excle.get_rows_number("test_api_004"))
    # print(hand_excle.get_rows_number('test_api_003'))
    # copyexcel(test_excel,report_result)
    # excel_data.get_success_result()
    print(excel_data.get_excel_data())