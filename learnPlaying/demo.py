#coding:utf-8

import xlrd
import openpyxl
import os
file  =  os.path.dirname(__file__) + '/APIcase.xlsx'
print(file)

# excel = xlrd.open_workbook(file)
# sheet = excel.sheet_by_index(0)
# cell = sheet.cell_value(1,1)
# print(cell,type(cell))

excel = openpyxl.load_workbook(file)
sheet = excel['Sheet1']
cell = sheet.cell(row=2,column=2).value
cell = sheet.cell(2,2).value
print(cell,type(cell))
